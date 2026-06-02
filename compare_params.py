"""
参数对比：原始版 vs 加MA250 vs 加MA250+连涨≥2天
用30年数据，买入=当日收盘，卖出=次日开盘，Top5，成本0.10%
"""
import os, sys, itertools
sys.path.insert(0, os.path.dirname(__file__))

import numpy as np, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import defaultdict
from backtest import read_day_file

DATA_DIR   = r"F:\OvernightStock\download\hsjday"
START_INT  = 19950101
END_INT    = 20260529
IDX_MIN    = 0.5
EXCESS_LO, EXCESS_HI = 2.0, 2.5
CHG_LO, CHG_HI = 3.0, 5.0
TOP_N, COST_PCT = 5, 0.10

# ── 指数技术指标 ──────────────────────────────────────────────────────────────
idf = read_day_file(os.path.join(DATA_DIR,"sh","lday","sh000001.day"))
idf = idf[idf["date"] >= 19930101].reset_index(drop=True)
idx_dates  = idf["date"].tolist()
idx_closes = idf["close"].tolist()

idx_chg_map = {idx_dates[i]: float((idx_closes[i]/idx_closes[i-1]-1)*100) if i>0 else 0.0
               for i in range(len(idx_dates))}

# 5MA
ma5_map = {}
for i,d in enumerate(idx_dates):
    if i<5: ma5_map[d]=None; continue
    ma5_map[d] = idx_closes[i] > sum(idx_closes[i-5:i])/5

# 250MA（年线）
ma250_map = {}
for i,d in enumerate(idx_dates):
    if i<250: ma250_map[d]=None; continue
    ma250_map[d] = idx_closes[i] > sum(idx_closes[i-250:i])/250

# 连续上涨天数
consec_map = {}
streak = 0
for i,d in enumerate(idx_dates):
    if i==0: consec_map[d]=0; continue
    chg = idx_chg_map.get(d,0)
    streak = streak+1 if chg>0 else 0
    consec_map[d] = streak

# ── 扫描股票（宽范围预加载）────────────────────────────────────────────────────
def load_stock(args):
    code, path = args
    try:
        df = read_day_file(path)
        if len(df)<25: return []
        df = df.set_index("date")
        close=df["close"]; open_=df["open"]; prev_close=close.shift(1)
        chg_pct=(close/prev_close-1)*100; next_open=open_.shift(-1)
        mask=((df.index>=START_INT)&(df.index<=END_INT)
              &next_open.notna()&prev_close.notna())
        rows=[]
        for date_int, row in df[mask].iterrows():
            chg=float(chg_pct.get(date_int,0))
            ic=idx_chg_map.get(date_int,0.0)
            if not (CHG_LO<=chg<=CHG_HI): continue
            if ic<IDX_MIN: continue
            excess=chg/ic if ic>0 else 0
            if not (EXCESS_LO<=excess<=EXCESS_HI): continue
            if ma5_map.get(date_int) is not True: continue
            bp=float(row["close"]); sp=float(next_open[date_int])
            if bp<=0 or sp<=0: continue
            net=(sp/bp-1)*100-COST_PCT
            rows.append({
                "date":    date_int,
                "code":    code,
                "year":    str(date_int)[:4],
                "chg":     round(chg,2),
                "ic":      round(ic,2),
                "net":     round(net,2),
                "wday":    __import__('datetime').date(int(str(date_int)[:4]),int(str(date_int)[4:6]),int(str(date_int)[6:])).weekday(),
                "ma250ok": ma250_map.get(date_int),
                "consec":  consec_map.get(date_int,0),
            })
        return rows
    except Exception: return []

print("加载沪市股票…")
paths=[(fn.replace(".day","")[2:], os.path.join(DATA_DIR,"sh","lday",fn))
       for fn in os.listdir(os.path.join(DATA_DIR,"sh","lday"))
       if fn.endswith(".day") and not fn[2:8].startswith("000")]

all_rows=[]
with ThreadPoolExecutor(max_workers=8) as ex:
    futs=[ex.submit(load_stock,p) for p in paths]
    for i,fut in enumerate(as_completed(futs)):
        all_rows.extend(fut.result())
        if (i+1)%2000==0: print(f"  {i+1}/{len(paths)}…")

# 每日Top5
by_date=defaultdict(list)
for r in all_rows: by_date[r["date"]].append(r)
base_trades=[]
for d in sorted(by_date):
    day=sorted(by_date[d],key=lambda x:x["chg"],reverse=True)[:TOP_N]
    base_trades.extend(day)
print(f"基础：{len(base_trades)} 笔")

# ── 三个参数版本 ──────────────────────────────────────────────────────────────
def run_version(trades, label, ma250_filter=False, consec_filter=False):
    filtered=[]
    for t in trades:
        if ma250_filter and t["ma250ok"] is not True: continue
        if consec_filter and t["consec"] < 2: continue
        filtered.append(t)
    if not filtered: return None
    rets=np.array([t["net"] for t in filtered])
    wins=(rets>0).sum()
    by_year=defaultdict(list)
    for t in filtered: by_year[t["year"]].append(t["net"])
    by_wday=defaultdict(list)
    for t in filtered: by_wday[t["wday"]].append(t["net"])
    return {
        "label": label, "n": len(filtered),
        "wr": round((rets>0).mean()*100,1),
        "avg": round(rets.mean(),3),
        "by_year": {y: (round(np.mean(v),3), round((np.array(v)>0).mean()*100,1), len(v))
                    for y,v in sorted(by_year.items())},
        "by_wday": {d: (round(np.mean(v),3), round((np.array(v)>0).mean()*100,1))
                    for d,v in by_wday.items()},
    }

v1 = run_version(base_trades, "原始（5MA+超额2-2.5x）")
v2 = run_version(base_trades, "+年线MA250保护",           ma250_filter=True)
v3 = run_version(base_trades, "+年线MA250+连涨≥2天",     ma250_filter=True, consec_filter=True)

WDAY_NAMES={0:"周一",1:"周二",2:"周三",3:"周四",4:"周五"}

print("\n=== 三版本对比 ===")
for v in [v1,v2,v3]:
    if v: print(f"  {v['label']:30s}  笔:{v['n']:6d}  胜率:{v['wr']}%  均收益:{v['avg']:+.3f}%")

print("\n=== 年度对比（V1 vs V3）===")
all_years = sorted({y for v in [v1,v3] if v for y in v["by_year"]})
for y in all_years:
    r1=v1["by_year"].get(y,("—","—",0)) if v1 else ("—","—",0)
    r3=v3["by_year"].get(y,("—","—",0)) if v3 else ("—","—",0)
    flag="✅" if isinstance(r3[0],float) and r3[0]>0 else "❌"
    print(f"  {y}  V1:{r1[0]:+.3f}%({r1[2]}笔)  V3:{r3[0] if isinstance(r3[0],float) else r3[0]:}%({r3[2]}笔)  {flag}")

# ── 写 Excel ──────────────────────────────────────────────────────────────────
GREEN="FF1B5E20"; RED="FFB71C1C"; GREY="FF424242"
HEADER_BG="FF1A237E"
thin=Side(style="thin",color="FFB0BEC5")
bdr=Border(left=thin,right=thin,top=thin,bottom=thin)

wb=openpyxl.load_workbook(r"F:\OvernightStock\test.xlsx")
sname="参数对比（30年）"
if sname in wb.sheetnames: del wb[sname]
ws=wb.create_sheet(sname,0)

for col,w in enumerate([30,8,8,8,10,10,10,10,10,10],1):
    ws.column_dimensions[chr(64+col)].width=w
ws.row_dimensions[1].height=28

ws.merge_cells("A1:J1")
c=ws["A1"]
c.value="三版本参数对比（30年·1995-2026·沪市·Top5·次日开盘卖出·成本0.10%）"
c.font=Font(bold=True,size=11,color="FFFFFFFF"); c.fill=PatternFill("solid",fgColor=HEADER_BG)
c.alignment=Alignment(horizontal="center",vertical="center")

versions=[v1,v2,v3]
v_colors=["FF1A237E","FF1B5E20","FF4A148C"]

# 总览行
for vi,(v,vc) in enumerate(zip(versions,v_colors)):
    if not v: continue
    row=vi+2
    ws.row_dimensions[row].height=20
    ws.merge_cells(f"A{row}:D{row}")
    c=ws.cell(row=row,column=1,value=v["label"])
    c.font=Font(bold=True,size=10,color="FFFFFFFF"); c.fill=PatternFill("solid",fgColor=vc)
    c.alignment=Alignment(horizontal="left",vertical="center")
    for col,val in enumerate([v["n"],f"{v['wr']}%",f"{v['avg']:+.3f}%","","",""],5):
        c=ws.cell(row=row,column=col,value=val)
        c.fill=PatternFill("solid",fgColor=vc); c.border=bdr
        c.font=Font(size=10,bold=True,color="FFFFFFFF")
        c.alignment=Alignment(horizontal="center",vertical="center")

# 年度对比表
row=6
ws.row_dimensions[row].height=18
for col,h in enumerate(["年份","V1笔","V1胜率","V1均收益","V2笔","V2胜率","V2均收益","V3笔","V3胜率","V3均收益"],1):
    c=ws.cell(row=row,column=col,value=h)
    c.font=Font(bold=True,color="FFFFFFFF",size=9); c.fill=PatternFill("solid",fgColor="FF37474F")
    c.alignment=Alignment(horizontal="center",vertical="center"); c.border=bdr
row+=1

for y in all_years:
    ws.row_dimensions[row].height=15
    r1=v1["by_year"].get(y) if v1 else None
    r2=v2["by_year"].get(y) if v2 else None
    r3=v3["by_year"].get(y) if v3 else None
    bg="FFE8F5E9" if (r3 and r3[0]>0) else ("FFFCE4EC" if (r3 and r3[0]<-0.1) else "FFFFFFFF")
    vals=[y]
    for rd in [r1,r2,r3]:
        if rd: vals+=[rd[2],f"{rd[1]}%",f"{rd[0]:+.3f}%"]
        else:  vals+=["—","—","—"]
    for col,v in enumerate(vals,1):
        c=ws.cell(row=row,column=col,value=v)
        c.fill=PatternFill("solid",fgColor=bg); c.border=bdr
        c.alignment=Alignment(horizontal="center",vertical="center")
        is_ret_col = col in (4,7,10)
        try:
            fv=float(v.replace("%","").replace("+","")) if isinstance(v,str) else v
            c.font=Font(size=9,color=GREEN if (is_ret_col and fv>0) else (RED if (is_ret_col and fv<0) else GREY))
        except: c.font=Font(size=9,color=GREY)
    row+=1

ws.freeze_panes="A7"
wb.save(r"F:\OvernightStock\test.xlsx")
print(f"\nExcel 已保存")
